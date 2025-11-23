import os
import matplotlib.pyplot as plt
from wordcloud import WordCloud, STOPWORDS
import nltk
from nltk.corpus import stopwords
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import pandas as pd

nltk.download('stopwords')

def visualize_and_save_results(topic_model, topic_to_inc, output_folder):
    print("\nGenerating intertopic distance map...")

    fig_topics = topic_model.visualize_topics(width=1200, height=800)
    pio.write_html(fig_topics, file=os.path.join(output_folder, 'intertopic_distance_map.html'), auto_open=False)
    print("Intertopic distance map saved as 'intertopic_distance_map.html'.")

    print("\nGenerating bar chart of top words per topic...")
    fig_barchart = topic_model.visualize_barchart(top_n_topics=50, n_words=10, width=1200, height=800)
    pio.write_html(fig_barchart, file=os.path.join(output_folder, 'top_words_per_topic.html'), auto_open=False)
    print("Bar chart of top words per topic saved as 'top_words_per_topic.html'.")

    with open(os.path.join(output_folder, 'topics_and_inc_numbers.txt'), 'w') as f:
        for topic, inc_numbers in topic_to_inc.items():
            f.write(f"Topic: {topic}\n")
            f.write(f"INC-numbers: {', '.join(map(str, inc_numbers))}\n\n")
    print("Topics and their associated INC-numbers saved as 'topics_and_inc_numbers.txt'.")

def preprocess_for_wordcloud(texts, forbidden_words):
    stopwords_english = set(STOPWORDS)
    stopwords_dutch = set(stopwords.words("dutch"))
    stopwords_combined = stopwords_english.union(stopwords_dutch).union(set(forbidden_words))
    
    def clean_text(text):
        if not isinstance(text, str):
            text = str(text)
        tokens = text.split()
        tokens = [token.lower() for token in tokens if token.isalpha() and token.lower() not in stopwords_combined]
        return ' '.join(tokens)
    
    preprocessed_texts = [clean_text(text) for text in texts]
    return ' '.join(preprocessed_texts)

def generate_wordcloud(texts, forbidden_words, output_folder):
    print("\nGenerating word cloud...")
    
    data = preprocess_for_wordcloud(texts, forbidden_words)

    print(f"Word cloud data: {data[:100]}...")  # Print the first 100 characters

    if not data.strip():
        print("No words available to generate the word cloud.")
        return

    wordcloud = WordCloud(
        background_color='white',
        stopwords=STOPWORDS,
        max_words=200,
        max_font_size=40,
        scale=3,
        random_state=1
    ).generate(data)

    plt.figure(figsize=(12, 12))
    plt.axis('off')
    plt.imshow(wordcloud, interpolation='bilinear')
    wordcloud_path = os.path.join(output_folder, 'wordcloud.png')
    plt.savefig(wordcloud_path)
    plt.close()
    print(f"Word cloud saved as '{wordcloud_path}'.")

from openpyxl.worksheet.datavalidation import DataValidation

def save_to_excel(topic_model, data, output_folder):
    topic_info = topic_model.get_topic_info()
    top_topics = topic_model.get_topic_freq()

    def clean_topic_names(topic_model, top_topics):
        names = []
        for topic in top_topics['Topic']:
            if topic == -1:
                names.append("Outliers")
            else:
                words = [word for word, _ in topic_model.get_topic(topic)]
                name = ", ".join(words[:5])
                names.append(name)
        return names

    top_topics['Name'] = clean_topic_names(topic_model, top_topics)

    output_file = os.path.join(output_folder, 'topic_modeling_report.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        topic_info.to_excel(writer, sheet_name='Topic Information', index=False)
        top_topics.to_excel(writer, sheet_name='Top Topics', index=False)

        # **Overige visualisatiebladen behouden**
        worksheet_sentiment = writer.book.create_sheet('Sentiment Distribution')
        img_sentiment = Image(os.path.join(output_folder, 'sentiment_distribution.png'))
        worksheet_sentiment.add_image(img_sentiment, 'A1')

        worksheet_visuals = writer.book.create_sheet('Visualizations')
        worksheet_visuals['A1'] = 'Intertopic Distance Map'
        worksheet_visuals['A2'] = 'Top Words Per Topic'
        worksheet_visuals['A3'] = 'Word Cloud'
        worksheet_visuals['B1'] = 'Link'
        worksheet_visuals['B2'] = 'Link'
        worksheet_visuals['B3'] = 'Image'
        worksheet_visuals['B1'].hyperlink = 'intertopic_distance_map.html'
        worksheet_visuals['B2'].hyperlink = 'top_words_per_topic.html'
        worksheet_visuals['B1'].style = 'Hyperlink'
        worksheet_visuals['B2'].style = 'Hyperlink'

        img_wordcloud = Image(os.path.join(output_folder, 'wordcloud.png'))
        worksheet_visuals.add_image(img_wordcloud, 'A3')

    writer.book.save(output_file)
    print(f"Excel-rapport gegenereerd: '{output_file}'")

def generate_html_report(topic_to_inc, topic_names, html_data_subset, output_folder):
    html_path = os.path.join(output_folder, "topic_report.html")

    html_content = """<!DOCTYPE html>
<html lang="nl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>NLP Topic Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { text-align: center; }
        .topic { margin-bottom: 10px; padding: 10px; border: 1px solid #ddd; border-radius: 5px; background: #f9f9f9; cursor: pointer; }
        .inc-table { display: none; width: 100%; border-collapse: collapse; margin-top: 10px; }
        .inc-table th, .inc-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        .inc-table th { background-color: #f2f2f2; cursor: pointer; }
        .bold { font-weight: bold; }
        input { margin-bottom: 10px; padding: 5px; width: 50%; }
    </style>
    <script>
        function toggleIncTable(id) {
            var table = document.getElementById(id);
            table.style.display = (table.style.display === "none") ? "table" : "none";
        }

        function filterTable(topicId, colIndex) {
            var input = document.getElementById("filter_" + topicId + "_" + colIndex).value.toLowerCase();
            var table = document.getElementById(topicId);
            var rows = table.getElementsByTagName("tr");

            for (var i = 1; i < rows.length; i++) {
                var cell = rows[i].getElementsByTagName("td")[colIndex];
                if (cell) {
                    if (cell.innerText.toLowerCase().includes(input)) {
                        rows[i].style.display = "";
                    } else {
                        rows[i].style.display = "none";
                    }
                }
            }
        }
    </script>
</head>
<body>
    <h1>Topic Report</h1>
    <p>Klik op een topic om de bijbehorende INC-nummers en details te bekijken.</p>
"""

    # **Kolomnamen instellen**
    column_names = ["nummer", "opened_at", "service_offering", "assignment_group", "category", "reassignment_count", "reopen_count"]

    # **Topics sorteren van groot naar klein**
    sorted_topics = sorted(topic_to_inc.items(), key=lambda item: len(item[1]), reverse=True)

    for topic, inc_numbers in sorted_topics:
        topic_name = topic_names.get(topic, "Onbekend")
        num_inc = len(inc_numbers)

        html_content += f'<div class="topic" onclick="toggleIncTable(\'inc_table_{topic}\')">'
        html_content += f'📂 Topic {topic}: <span class="bold">({num_inc})</span> {topic_name}</div>\n'

        # **Tabel met filterbare invoervelden**
        html_content += f'<table id="inc_table_{topic}" class="inc-table">\n'
        html_content += "  <tr>\n"
        for col_index, col_name in enumerate(column_names):
            html_content += f'    <th>{col_name}<br><input type="text" id="filter_inc_table_{topic}_{col_index}" onkeyup="filterTable(\'inc_table_{topic}\', {col_index})" placeholder="Filter..."></th>\n'
        html_content += "  </tr>\n"

        # **Data invullen**
        for inc in inc_numbers:
            row = html_data_subset[html_data_subset['nummer'] == inc].iloc[0].to_dict()
            html_content += "  <tr>\n"
            for col in column_names:
                html_content += f'    <td>{row[col]}</td>\n'
            html_content += "  </tr>\n"

        html_content += "</table>\n"

    html_content += """
</body>
</html>
"""

    # **HTML-bestand opslaan**
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"✅ HTML-report gegenereerd: {html_path}")